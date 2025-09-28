# 📁 boats/admin.py - ПОЛНАЯ ИСПРАВЛЕННАЯ ВЕРСИЯ
# 🛥️ Максимально детализированная админка по аналогии с автомобилями
# ✅ УНИФИКАЦИЯ: Все функции ProductAdmin адаптированы для BoatProduct

import os
import logging
from django import forms
from django.contrib import admin
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.contrib import messages
from django.urls import path
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.core.exceptions import ValidationError

from .models import BoatCategory, BoatProduct, BoatProductImage, BoatCatalogDescription

logger = logging.getLogger(__name__)


class BoatExcelImportForm(forms.Form):
    """📊 Форма импорта Excel для лодок"""
    excel_file = forms.FileField(
        label="📊 Excel файл с данными лодок",
        help_text="Формат: .xlsx, .xls. Колонки: A-Категория, B-Название, C-Длина(см), D-Ширина(см), E-Цена",
        widget=forms.FileInput(attrs={'accept': '.xlsx,.xls', 'class': 'form-control-file'})
    )
    images_zip = forms.FileField(
        label="🖼️ ZIP архив с изображениями",
        required=False,
        help_text="Необязательно. ZIP файл с изображениями лодок",
        widget=forms.FileInput(attrs={'accept': '.zip', 'class': 'form-control-file'})
    )


class BoatCategoryAdminForm(forms.ModelForm):
    """📋 Форма категории лодок с валидацией"""

    class Meta:
        model = BoatCategory
        fields = '__all__'

    def clean(self):
        """🔧 Валидация данных категории"""
        cleaned_data = super().clean()
        category_name = cleaned_data.get('category_name')

        if category_name and len(category_name) < 3:
            raise ValidationError("Название категории должно содержать минимум 3 символа")

        return cleaned_data


class BoatProductImageInline(admin.TabularInline):
    """🖼️ Инлайн для изображений лодочных товаров с детальной настройкой"""
    model = BoatProductImage
    extra = 1
    max_num = 10
    fields = ['image', 'alt_text', 'is_main', 'display_order', 'get_image_preview']
    readonly_fields = ['get_image_preview', 'created_at']

    def get_image_preview(self, obj):
        """🖼️ Превью изображения в инлайне"""
        if obj.image:
            return format_html(
                '<img src="{}" style="width: 60px; height: 60px; object-fit: cover; border-radius: 5px;" />',
                obj.image.url
            )
        return "Нет изображения"

    get_image_preview.short_description = "Превью"


@admin.register(BoatCategory)
class BoatCategoryAdmin(admin.ModelAdmin):
    """🛥️ Детализированная админка категорий лодок"""

    form = BoatCategoryAdminForm

    list_display = [
        'get_category_preview',
        'category_name',
        'get_products_count',
        'display_order',
        'created_at',
        'get_seo_status'
    ]

    list_display_links = ['get_category_preview', 'category_name']
    list_filter = ['created_at']
    search_fields = ['category_name', 'description', 'page_title', 'meta_title']
    list_editable = ['display_order']
    list_per_page = 25

    prepopulated_fields = {'slug': ('category_name',)}

    readonly_fields = [
        'created_at', 'updated_at', 'get_image_preview', 'get_meta_title_length',
        'get_meta_description_length', 'get_google_preview'
    ]

    fieldsets = (
        ('🛥️ Основная информация', {
            'fields': ('category_name', 'slug', 'category_image', 'get_image_preview', 'display_order'),
            'description': '🏷️ Базовая информация о категории лодок'
        }),
        ('📝 Контент категории', {
            'fields': ('description', 'additional_content'),
            'classes': ('wide',),
            'description': '✍️ Описание и дополнительный контент (YouTube видео)'
        }),
        ('🔍 SEO-настройки', {
            'fields': (
                'page_title',
                ('meta_title', 'get_meta_title_length'),
                ('meta_description', 'get_meta_description_length'),
                'get_google_preview'
            ),
            'description': '🎯 Оптимизация для поисковых систем'
        }),
        ('⚙️ Настройки отображения', {
            'fields': ('is_active',),
            'classes': ('collapse',),
            'description': '🔧 Видимость категории'
        }),
        ('📅 Служебная информация', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

    actions = ['activate_categories', 'deactivate_categories', 'optimize_seo']

    def get_category_preview(self, obj):
        """🖼️ Превью изображения категории"""
        if obj.category_image:
            return format_html(
                '<img src="{}" style="width: 50px; height: 50px; object-fit: cover; border-radius: 5px; border: 2px solid #f39c12;" title="{}">',
                obj.category_image.url,
                obj.category_name
            )
        return "📷"

    get_category_preview.short_description = "Изображение"

    def get_products_count(self, obj):
        """📊 Количество товаров в категории"""
        count = obj.get_products_count()
        if count > 0:
            return format_html(
                '<span style="color: green; font-weight: bold;">📦 {}</span>',
                count
            )
        return format_html('<span style="color: gray;">пусто</span>')

    get_products_count.short_description = "Товаров"

    def get_seo_status(self, obj):
        """🔍 Статус SEO оптимизации"""
        seo_fields = [obj.page_title, obj.meta_title, obj.meta_description]
        filled_fields = sum(1 for field in seo_fields if field)

        if filled_fields >= 2:
            return format_html('<span style="color: green;">✅ Настроено</span>')
        elif filled_fields == 1:
            return format_html('<span style="color: orange;">⚠️ Частично</span>')
        return format_html('<span style="color: red;">❌ Не настроено</span>')

    get_seo_status.short_description = "SEO"

    def get_image_preview(self, obj):
        """🖼️ Предпросмотр изображения категории"""
        if obj.category_image:
            return format_html(
                '<img src="{}" style="max-width: 200px; max-height: 150px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);" />',
                obj.category_image.url
            )
        return "❌ Изображение не загружено"

    get_image_preview.short_description = "Предпросмотр изображения"

    def get_meta_title_length(self, obj):
        """📏 Длина мета-заголовка"""
        if obj.meta_title:
            length = len(obj.meta_title)
            color = "green" if length <= 60 else "red"
            return format_html('<span style="color: {}; font-weight: bold;">{}/60</span>', color, length)
        return format_html('<span style="color: gray;">0/60</span>')

    get_meta_title_length.short_description = "Длина"

    def get_meta_description_length(self, obj):
        """📏 Длина мета-описания"""
        if obj.meta_description:
            length = len(obj.meta_description)
            color = "green" if length <= 160 else "red"
            return format_html('<span style="color: {}; font-weight: bold;">{}/160</span>', color, length)
        return format_html('<span style="color: gray;">0/160</span>')

    get_meta_description_length.short_description = "Длина"

    def get_google_preview(self, obj):
        """🔍 Предпросмотр в стиле Google"""
        title = obj.meta_title or obj.page_title or obj.category_name
        description = obj.meta_description or "Описание не заполнено"

        return format_html(
            '<div style="border: 1px solid #dadce0; border-radius: 8px; padding: 12px; max-width: 500px; font-family: Arial, sans-serif; margin: 10px 0;">'
            '<div style="color: #1a0dab; font-size: 18px; line-height: 1.3; margin-bottom: 4px;">{}</div>'
            '<div style="color: #006621; font-size: 14px; margin-bottom: 4px;">https://автоковрик.бай/boats/{}/</div>'
            '<div style="color: #545454; font-size: 14px; line-height: 1.4;">{}</div>'
            '</div>',
            title,
            obj.slug or 'category-slug',
            description
        )

    get_google_preview.short_description = "Предпросмотр Google"

    def activate_categories(self, request, queryset):
        """✅ Активировать выбранные категории"""
        updated = queryset.update(is_active=True)
        self.message_user(request, f"✅ Активировано категорий: {updated}")

    def deactivate_categories(self, request, queryset):
        """❌ Деактивировать выбранные категории"""
        updated = queryset.update(is_active=False)
        self.message_user(request, f"❌ Деактивировано категорий: {updated}")

    def optimize_seo(self, request, queryset):
        """🔍 Автоматическая SEO оптимизация"""
        optimized = 0
        for category in queryset:
            changed = False
            if not category.page_title:
                category.page_title = category.category_name
                changed = True
            if not category.meta_title:
                category.meta_title = f"ЭВА коврики для лодок {category.category_name}"[:60]
                changed = True
            if not category.meta_description:
                category.meta_description = f"Качественные ЭВА коврики для лодок {category.category_name.lower()}. Защита дна лодки, выбор цвета, доставка по Беларуси."[:160]
                changed = True
            if changed:
                category.save()
                optimized += 1
        self.message_user(request, f"🔍 SEO оптимизировано для {optimized} категорий")

    activate_categories.short_description = "✅ Активировать категории"
    deactivate_categories.short_description = "❌ Деактивировать категории"
    optimize_seo.short_description = "🔍 Оптимизировать SEO"


@admin.register(BoatProduct)
class BoatProductAdmin(admin.ModelAdmin):
    """🛥️ ПОЛНАЯ ДЕТАЛИЗИРОВАННАЯ админка товаров лодок"""

    # 🔗 Подключаем изображения
    inlines = [BoatProductImageInline]

    list_display = [
        'get_main_image_preview',
        'product_name',
        'product_sku',
        'category',
        'get_price_display',
        'get_dimensions_badge',
        'has_main_image_status',
        'newest_product',
        'created_at'
    ]

    list_display_links = ['get_main_image_preview', 'product_name']

    list_filter = [
        'category',
        'newest_product',
        'created_at',
        'boat_mat_length',
        'boat_mat_width'
    ]

    search_fields = [
        'product_name',
        'product_sku',
        'product_desription',
        'boat_mat_length',
        'boat_mat_width'
    ]

    list_editable = ['newest_product']
    list_per_page = 25

    prepopulated_fields = {'slug': ('product_name',)}

    readonly_fields = [
        'product_sku',  # Автогенерируется
        'created_at',
        'updated_at',
        'get_main_image_large'
    ]

    # 📝 Детализированные fieldsets
    fieldsets = (
        ('🛍️ Основная информация', {
            'fields': ('product_sku', 'product_name', 'slug', 'category', 'price')
        }),
        ('🛥️ Размеры лодочного коврика', {
            'fields': ('boat_mat_length', 'boat_mat_width'),
            'description': '📏 Размеры коврика в сантиметрах для лодки.',
        }),
        ('📝 Описание товара', {
            'fields': ('product_desription',),
            'classes': ('wide',)
        }),
        ('🔍 SEO-настройки', {
            'fields': ('page_title', 'meta_description'),
            'classes': ('collapse',),
            'description': 'Настройки для поисковых систем'
        }),
        ('⚙️ Настройки отображения', {
            'fields': ('newest_product',)
        }),
        ('🖼️ Главное изображение', {
            'fields': ('get_main_image_large',),
            'classes': ('collapse',)
        }),
        ('📊 Служебная информация', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    # 🎯 Массовые действия
    actions = [
        'mark_as_new',
        'mark_as_regular',
        'set_first_image_as_main',
        'generate_missing_slugs',
        'optimize_boat_seo',
        'export_boats_excel'
    ]

    def get_main_image_preview(self, obj):
        """🖼️ Предпросмотр главного изображения товара"""
        main_image = obj.images.filter(is_main=True).first()
        if main_image and main_image.image:
            return format_html(
                '<img src="{}" style="width: 60px; height: 60px; object-fit: cover; border-radius: 5px; border: 2px solid #f39c12;" title="{}">',
                main_image.image.url,
                obj.product_name
            )

        # Если нет главного, берем первое доступное
        first_image = obj.images.first()
        if first_image and first_image.image:
            return format_html(
                '<img src="{}" style="width: 60px; height: 60px; object-fit: cover; border-radius: 5px; border: 1px solid #ddd;" title="{}">',
                first_image.image.url,
                obj.product_name
            )

        return "📷"

    get_main_image_preview.short_description = "Фото"

    def get_main_image_large(self, obj):
        """🖼️ Большое изображение для детальной страницы"""
        main_image = obj.get_main_image()
        if main_image and main_image.image:
            return format_html(
                '<img src="{}" style="max-width: 300px; max-height: 300px; border-radius: 10px;" />',
                main_image.image.url
            )
        return "Нет главного изображения"

    get_main_image_large.short_description = "Главное изображение"

    def get_price_display(self, obj):
        """💰 Отображение цены в удобном формате"""
        if obj.price:
            return format_html(
                '<span style="color: green; font-weight: bold;">💰 {} руб.</span>',
                f"{obj.price:,}".replace(',', ' ')
            )
        return format_html('<span style="color: gray;">Не указана</span>')

    get_price_display.short_description = "Цена"
    get_price_display.admin_order_field = "price"

    def get_dimensions_badge(self, obj):
        """📏 Красивое отображение размеров коврика"""
        if obj.boat_mat_length and obj.boat_mat_width:
            area = round(obj.boat_mat_length * obj.boat_mat_width / 10000, 2)
            return format_html(
                '<span style="background: #e3f2fd; padding: 2px 6px; border-radius: 3px; font-size: 12px;" title="Площадь: {} м²">'
                '📏 {}×{}см'
                '</span>',
                area,
                obj.boat_mat_length,
                obj.boat_mat_width
            )
        elif obj.boat_mat_length:
            return format_html(
                '<span style="background: #fff3e0; padding: 2px 6px; border-radius: 3px; font-size: 12px;">'
                '📏 {}см (длина)'
                '</span>',
                obj.boat_mat_length
            )
        elif obj.boat_mat_width:
            return format_html(
                '<span style="background: #fff3e0; padding: 2px 6px; border-radius: 3px; font-size: 12px;">'
                '📏 {}см (ширина)'
                '</span>',
                obj.boat_mat_width
            )
        else:
            return format_html('<span style="color: orange;">⚠️ Не указаны</span>')

    get_dimensions_badge.short_description = "Размеры коврика"
    get_dimensions_badge.admin_order_field = "boat_mat_length"

    def has_main_image_status(self, obj):
        """🖼️ Статус главного изображения"""
        if obj.images.filter(is_main=True).exists():
            return format_html('<span style="color: green;">✅ Есть</span>')
        elif obj.images.exists():
            return format_html('<span style="color: orange;">⚠️ Не выбрано</span>')
        return format_html('<span style="color: red;">❌ Нет фото</span>')

    has_main_image_status.short_description = "Главное фото"

    # 🎯 МАССОВЫЕ ДЕЙСТВИЯ

    def mark_as_new(self, request, queryset):
        """🆕 Отметить как новые товары"""
        updated = queryset.update(newest_product=True)
        self.message_user(request, f"🆕 Отмечено как новые: {updated} товаров")

    def mark_as_regular(self, request, queryset):
        """📦 Убрать отметку 'новый'"""
        updated = queryset.update(newest_product=False)
        self.message_user(request, f"📦 Убрана отметка 'новый': {updated} товаров")

    def set_first_image_as_main(self, request, queryset):
        """🖼️ Установить первое фото как главное"""
        updated = 0
        for product in queryset:
            # Сбрасываем все главные изображения
            product.images.update(is_main=False)
            # Устанавливаем первое как главное
            first_image = product.images.first()
            if first_image:
                first_image.is_main = True
                first_image.save()
                updated += 1

        self.message_user(request, f"🖼️ Установлено главное фото для {updated} товаров")

    def generate_missing_slugs(self, request, queryset):
        """🔗 Сгенерировать отсутствующие slug"""
        from django.utils.text import slugify
        updated = 0
        for product in queryset.filter(slug__isnull=True):
            product.slug = slugify(product.product_name, allow_unicode=True)
            product.save()
            updated += 1

        self.message_user(request, f"🔗 Сгенерировано slug для {updated} товаров")

    def optimize_boat_seo(self, request, queryset):
        """🔍 Автоматическая SEO оптимизация для лодок"""
        optimized = 0
        for product in queryset:
            changed = False
            if not product.page_title:
                dimensions = f" {product.get_mat_dimensions()}" if product.boat_mat_length else ""
                product.page_title = f"🛥️ {product.product_name}{dimensions} - Купить лодочный коврик"
                changed = True

            if not product.meta_description:
                category_name = product.category.category_name
                dimensions_text = f" размером {product.get_mat_dimensions()}" if product.boat_mat_length else ""
                product.meta_description = f"Лодочный коврик {product.product_name} для {category_name}{dimensions_text}. EVA материал, точные лекала. Доставка по Беларуси."[
                                           :160]
                changed = True

            if changed:
                product.save()
                optimized += 1

        self.message_user(request, f"🔍 SEO оптимизировано для {optimized} товаров")

    def export_boats_excel(self, request, queryset):
        """📊 Экспорт лодок в Excel"""
        try:
            import openpyxl
            from django.http import HttpResponse
            from datetime import datetime

            # Создаем книгу Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Лодочные товары"

            # Заголовки
            headers = [
                'Артикул', 'Название', 'Категория', 'Цена',
                'Длина (см)', 'Ширина (см)', 'Площадь (м²)',
                'Новинка', 'Дата создания', 'URL'
            ]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Данные товаров
            for row, product in enumerate(queryset, 2):
                area = ""
                if product.boat_mat_length and product.boat_mat_width:
                    area = round(product.boat_mat_length * product.boat_mat_width / 10000, 2)

                ws.cell(row=row, column=1, value=product.product_sku or "")
                ws.cell(row=row, column=2, value=product.product_name)
                ws.cell(row=row, column=3, value=product.category.category_name)
                ws.cell(row=row, column=4, value=product.price or 0)
                ws.cell(row=row, column=5, value=product.boat_mat_length or "")
                ws.cell(row=row, column=6, value=product.boat_mat_width or "")
                ws.cell(row=row, column=7, value=area)
                ws.cell(row=row, column=8, value="Да" if product.newest_product else "Нет")
                ws.cell(row=row, column=9, value=product.created_at.strftime("%d.%m.%Y"))
                ws.cell(row=row, column=10, value=f"/boats/{product.slug}/")

            # Настройка ответа
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"boats_export_{timestamp}.xlsx"

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)
            return response

        except ImportError:
            self.message_user(request, "❌ Для экспорта нужно установить openpyxl: pip install openpyxl",
                              level=messages.ERROR)
        except Exception as e:
            self.message_user(request, f"❌ Ошибка экспорта: {str(e)}", level=messages.ERROR)

    # Подписи для actions
    mark_as_new.short_description = "🆕 Отметить как новые товары"
    mark_as_regular.short_description = "📦 Убрать отметку 'новый'"
    set_first_image_as_main.short_description = "🖼️ Установить первое фото как главное"
    generate_missing_slugs.short_description = "🔗 Сгенерировать отсутствующие slug"
    optimize_boat_seo.short_description = "🔍 Оптимизировать SEO"
    export_boats_excel.short_description = "📊 Экспорт в Excel"

    # 📊 Дополнительные URL для импорта
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-boats/',
                self.admin_site.admin_view(self.import_boats_view),
                name='boats_boatproduct_import'
            ),
        ]
        return custom_urls + urls

    def import_boats_view(self, request):
        """📊 Страница импорта Excel для лодок"""
        if request.method == 'POST':
            form = BoatExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                return self._process_boats_import(request, form)
        else:
            form = BoatExcelImportForm()

        context = {
            'form': form,
            'title': '📊 Импорт лодок из Excel',
            'subtitle': 'Загрузка категорий и товаров лодок',
            'opts': self.model._meta,
            'app_label': self.model._meta.app_label,
            'has_change_permission': True,
            'has_view_permission': True,
        }

        return render(request, 'admin/boats/import_boats.html', context)

    def _process_boats_import(self, request, form):
        """🔄 Обработка загруженного Excel файла"""
        try:
            excel_file = request.FILES['excel_file']
            result = self._process_boats_excel(excel_file)

            if result['success']:
                success_msg = f"✅ Импорт завершен! Обработано: {result['count']} товаров"
                messages.success(request, success_msg)

                for detail in result.get('details', []):
                    messages.info(request, detail)

                for warning in result.get('warnings', []):
                    messages.warning(request, warning)
            else:
                messages.error(request, f"❌ Ошибка импорта: {result['error']}")

        except Exception as e:
            logger.exception("Ошибка при импорте лодок")
            messages.error(request, f"❌ Ошибка при обработке файла: {str(e)}")

        return HttpResponseRedirect("../")

    def _process_boats_excel(self, excel_file):
        """📊 Парсинг Excel файла с лодками"""
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            count = 0
            details = []
            warnings = []

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if not any(row):
                    continue

                try:
                    category_name = str(row[0]).strip() if row[0] else ""
                    product_name = str(row[1]).strip() if row[1] else ""
                    length = int(row[2]) if row[2] and str(row[2]).isdigit() else None
                    width = int(row[3]) if row[3] and str(row[3]).isdigit() else None
                    price = int(float(str(row[4]))) if row[4] else 0

                    if not category_name or not product_name:
                        warnings.append(f"Строка {row_num}: Пропущена - нет названия")
                        continue

                    # Создаем/получаем категорию
                    category, created = BoatCategory.objects.get_or_create(
                        category_name=category_name,
                        defaults={'is_active': True}
                    )

                    # Создаем товар
                    boat_product = BoatProduct.objects.create(
                        product_name=product_name,
                        category=category,
                        price=price,
                        boat_mat_length=length,
                        boat_mat_width=width,
                        product_desription=f"Лодочный коврик {product_name}",
                        newest_product=False
                    )

                    count += 1
                    dimensions = f" ({length}×{width}см)" if length and width else ""
                    details.append(f"✅ {product_name}{dimensions}")

                except Exception as e:
                    warnings.append(f"Строка {row_num}: {str(e)}")
                    continue

            return {
                'success': True,
                'count': count,
                'details': details,
                'warnings': warnings
            }

        except Exception as e:
            return {'success': False, 'error': str(e)}


# 🛥️ НОВАЯ АДМИНКА: BoatCatalogDescription (синглтон)
@admin.register(BoatCatalogDescription)
class BoatCatalogDescriptionAdmin(admin.ModelAdmin):
    """📝 Админка для описания каталога лодочных ковриков (только один экземпляр)"""

    # 🚫 Синглтон логика в админке
    def has_add_permission(self, request):
        """🚫 Запретить создание новых записей, если уже есть описание"""
        return not BoatCatalogDescription.objects.exists()

    def has_delete_permission(self, request, obj=None):
        """⚠️ Разрешить удаление (чтобы можно было пересоздать при необходимости)"""
        return True

    def changelist_view(self, request, extra_context=None):
        """📝 Если нет записи, перенаправляем на создание"""
        if not BoatCatalogDescription.objects.exists():
            return self.add_view(request)
        return super().changelist_view(request, extra_context)

    # 🎨 Группировка полей в админке
    fieldsets = (
        ('🛥️ Описание каталога лодочных ковриков', {
            'fields': ('title', 'description'),
            'description': 'Заголовок и основное описание каталога лодочных ковриков'
        }),
        ('🎬 Дополнительный контент', {
            'fields': ('additional_content',),
            'classes': ('collapse',),
            'description': 'YouTube видео или дополнительный HTML контент'
        }),
        ('🔍 SEO', {
            'fields': ('meta_description',),
            'classes': ('collapse',),
            'description': 'Мета-описание для поисковых систем'
        }),
    )


# 🎨 Настройка заголовков админки лодок
admin.site.site_header = "🛥️🚗 Автоковрики - Админ-панель (Авто + Лодки)"
admin.site.site_title = "Автоковрики"
admin.site.index_title = "Управление интернет-магазином автомобильных и лодочных ковриков"