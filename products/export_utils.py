# 📁 products/export_utils.py
# 📊 Генерация Excel файла для экспорта товаров и категорий
# 🔄 Обратная операция к import_utils.py

import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Optional
from io import BytesIO
from django.db.models import Q

from .models import Category, Product

logger = logging.getLogger(__name__)

# 📋 Заголовки колонок Excel (как в импорте)
EXCEL_HEADERS = [
    'Идентификатор',  # A: Категория (1.BMW) или SKU товара (10001)
    'Название',  # B: Название категории или товара
    'Title страницы',  # C: SEO заголовок
    'Цена',  # D: Цена (только для товаров)
    'Описание',  # E: Описание с HTML
    'Мета-описание',  # F: SEO описание
    'Изображение'  # G: Имя файла изображения
]


def generate_excel_export() -> BytesIO:
    """
    🚀 ГЛАВНАЯ ФУНКЦИЯ: Генерация Excel файла с товарами и категориями

    Создает файл в том же формате, что принимает импорт:
    - Категория → все её товары → следующая категория → её товары
    - Сохраняет только имена файлов изображений (без путей)

    Returns:
        BytesIO: Excel файл готовый для скачивания
    """
    try:
        logger.info("🚀 Начинаем генерацию Excel экспорта")

        # 📊 Создаем новую Excel книгу
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Товары и категории"

        # 🎨 Настраиваем стили
        _setup_excel_styles(worksheet)

        # 📝 Добавляем заголовки
        _add_headers(worksheet)

        # 📊 Получаем данные из БД
        export_data = _collect_export_data()

        # ✍️ Заполняем данные
        row_number = _fill_excel_data(worksheet, export_data)

        # 📏 Автоширина колонок
        _auto_resize_columns(worksheet)

        # 💾 Сохраняем в память
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info(f"✅ Excel экспорт сгенерирован: {row_number - 1} строк данных")
        return excel_buffer

    except Exception as e:
        logger.error(f"❌ Ошибка генерации Excel экспорта: {e}", exc_info=True)
        raise


def _setup_excel_styles(worksheet):
    """🎨 Настройка стилей Excel"""
    # 📏 Устанавливаем ширину колонок по умолчанию
    column_widths = {
        'A': 15,  # Идентификатор
        'B': 30,  # Название
        'C': 25,  # Title
        'D': 10,  # Цена
        'E': 40,  # Описание
        'F': 30,  # Мета-описание
        'G': 20  # Изображение
    }

    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width


def _add_headers(worksheet):
    """📝 Добавление заголовков в Excel"""
    # 🎨 Стиль заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # ✍️ Записываем заголовки
    for col_num, header in enumerate(EXCEL_HEADERS, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def _collect_export_data() -> List[Dict]:
    """
    📊 Сбор данных для экспорта из БД

    Порядок: категория → все её товары → следующая категория

    Returns:
        List[Dict]: Список строк для Excel
    """
    try:
        export_rows = []

        # 📂 Получаем все активные категории с сортировкой
        categories = Category.objects.filter(is_active=True).order_by('display_order', 'category_name')

        logger.info(f"📂 Найдено категорий для экспорта: {categories.count()}")

        for category in categories:
            # 🏷️ Добавляем строку категории
            category_row = _build_category_row(category)
            export_rows.append(category_row)

            # 🛍️ Получаем товары этой категории
            products = Product.objects.filter(category=category).order_by('product_name')

            logger.debug(f"📦 Категория {category.category_name}: {products.count()} товаров")

            # 🛍️ Добавляем строки товаров
            for product in products:
                product_row = _build_product_row(product)
                export_rows.append(product_row)

        logger.info(f"📊 Подготовлено строк для экспорта: {len(export_rows)}")
        return export_rows

    except Exception as e:
        logger.error(f"❌ Ошибка сбора данных: {e}")
        raise


def _build_category_row(category: Category) -> Dict:
    """
    📂 Формирование строки категории для Excel

    Args:
        category: Объект категории

    Returns:
        Dict: Данные строки категории
    """
    try:
        # 🔢 Формируем идентификатор категории (SKU.Название)
        category_identifier = f"{category.category_sku}.{category.category_name}"

        # 🖼️ Получаем имя файла изображения (без пути)
        image_name = ""
        if category.category_image:
            import os
            image_name = os.path.basename(category.category_image.name)

        # 📝 Очищаем HTML из описания (берем только текст)
        description = _clean_html_content(category.description)

        return {
            'identifier': category_identifier,
            'name': category.category_name,
            'title': category.page_title or category.category_name,
            'price': '',  # 💰 У категорий нет цены
            'description': description,
            'meta_description': category.meta_description or '',
            'image': image_name
        }

    except Exception as e:
        logger.error(f"❌ Ошибка формирования строки категории {category.category_name}: {e}")
        raise


def _build_product_row(product: Product) -> Dict:
    """
    🛍️ Формирование строки товара для Excel

    Args:
        product: Объект товара

    Returns:
        Dict: Данные строки товара
    """
    try:
        # 🔢 SKU товара или автогенерация
        product_sku = product.product_sku
        if not product_sku:
            # 🆕 Генерируем SKU по формуле category_sku * 10000 + 1
            category_sku = product.category.category_sku or 1
            product_sku = str(category_sku * 10000 + 1)
            logger.warning(f"⚠️ Сгенерирован SKU для товара {product.product_name}: {product_sku}")

        # 🖼️ Получаем главное изображение (только имя файла)
        image_name = ""
        main_image = product.get_main_image()
        if main_image and main_image.image:
            import os
            image_name = os.path.basename(main_image.image.name)

        # 📝 Очищаем HTML из описания
        description = _clean_html_content(product.product_desription)

        # 💰 Цена товара (базовая цена)
        price = product.price or 0

        return {
            'identifier': str(product_sku),
            'name': product.product_name,
            'title': product.page_title or '',
            'price': price,
            'description': description,
            'meta_description': product.meta_description or '',
            'image': image_name
        }

    except Exception as e:
        logger.error(f"❌ Ошибка формирования строки товара {product.product_name}: {e}")
        raise


def _clean_html_content(html_content: Optional[str]) -> str:
    """
    🧹 Очистка HTML контента для Excel

    Удаляет HTML теги, оставляет только текст

    Args:
        html_content: HTML контент

    Returns:
        str: Очищенный текст
    """
    if not html_content:
        return ""

    try:
        # 🧹 Простая очистка HTML тегов
        import re

        # Удаляем HTML теги
        clean_text = re.sub(r'<[^>]+>', '', str(html_content))

        # Убираем лишние пробелы и переносы
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()

        # 📏 Ограничиваем длину для Excel (максимум 1000 символов)
        if len(clean_text) > 1000:
            clean_text = clean_text[:997] + "..."

        return clean_text

    except Exception as e:
        logger.warning(f"⚠️ Ошибка очистки HTML: {e}")
        return str(html_content)[:100] if html_content else ""


def _fill_excel_data(worksheet, export_data: List[Dict]) -> int:
    """
    ✍️ Заполнение Excel данными

    Args:
        worksheet: Лист Excel
        export_data: Данные для записи

    Returns:
        int: Номер последней заполненной строки
    """
    try:
        current_row = 2  # 📝 Начинаем со второй строки (после заголовков)

        for row_data in export_data:
            # ✍️ Заполняем колонки
            worksheet.cell(row=current_row, column=1, value=row_data['identifier'])
            worksheet.cell(row=current_row, column=2, value=row_data['name'])
            worksheet.cell(row=current_row, column=3, value=row_data['title'])
            worksheet.cell(row=current_row, column=4, value=row_data['price'])
            worksheet.cell(row=current_row, column=5, value=row_data['description'])
            worksheet.cell(row=current_row, column=6, value=row_data['meta_description'])
            worksheet.cell(row=current_row, column=7, value=row_data['image'])

            current_row += 1

        logger.info(f"✅ Заполнено строк данных: {current_row - 2}")
        return current_row

    except Exception as e:
        logger.error(f"❌ Ошибка заполнения Excel: {e}")
        raise


def _auto_resize_columns(worksheet):
    """📏 Автоматическая настройка ширины колонок"""
    try:
        # 📊 Устанавливаем оптимальную ширину для каждой колонки
        optimal_widths = {
            'A': 18,  # Идентификатор (1.BMW, 10001)
            'B': 35,  # Название
            'C': 25,  # Title страницы
            'D': 12,  # Цена
            'E': 50,  # Описание (самая широкая)
            'F': 35,  # Мета-описание
            'G': 25  # Изображение
        }

        for col_letter, width in optimal_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        logger.debug("📏 Автоширина колонок настроена")

    except Exception as e:
        logger.warning(f"⚠️ Ошибка настройки ширины колонок: {e}")


def get_export_statistics() -> Dict:
    """
    📊 Получение статистики для экспорта (для отображения в UI)

    Returns:
        Dict: Статистика экспорта
    """
    try:
        # 📂 Считаем активные категории
        active_categories = Category.objects.filter(is_active=True).count()

        # 🛍️ Считаем товары в активных категориях
        products_in_active_categories = Product.objects.filter(
            category__is_active=True
        ).count()

        # 🖼️ Считаем товары с главными изображениями
        products_with_images = Product.objects.filter(
            category__is_active=True,
            product_images__is_main=True
        ).count()

        # 📂 Считаем категории с изображениями
        categories_with_images = Category.objects.filter(
            is_active=True,
            category_image__isnull=False
        ).exclude(category_image='').count()

        return {
            'total_categories': active_categories,
            'total_products': products_in_active_categories,
            'products_with_images': products_with_images,
            'categories_with_images': categories_with_images,
            'estimated_rows': active_categories + products_in_active_categories,
        }

    except Exception as e:
        logger.error(f"❌ Ошибка получения статистики: {e}")
        return {
            'total_categories': 0,
            'total_products': 0,
            'products_with_images': 0,
            'categories_with_images': 0,
            'estimated_rows': 0,
            'error': str(e)
        }

# 🎯 ОСНОВНЫЕ ФУНКЦИИ ЭТОГО ФАЙЛА:
#
# ✅ generate_excel_export() - главная функция генерации
# ✅ _collect_export_data() - сбор данных из БД
# ✅ _build_category_row() - формирование строки категории
# ✅ _build_product_row() - формирование строки товара
# ✅ _clean_html_content() - очистка HTML для Excel
# ✅ get_export_statistics() - статистика для UI
#
# 🔄 ЛОГИКА ЭКСПОРТА:
# 1. Получаем все активные категории
# 2. Для каждой категории: категория → все её товары
# 3. Генерируем SKU если отсутствует
# 4. Извлекаем только имена файлов изображений
# 5. Очищаем HTML от тегов
# 6. Создаем Excel и отдаем как BytesIO