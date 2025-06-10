# 📁 products/import_processor.py - ИСПРАВЛЕННАЯ ВЕРСИЯ
# 🔄 Основной процессор для импорта данных из Excel файлов - БЕЗ ОШИБКИ 'list' object is not callable

import os
import logging
from typing import List, Dict, Any, Optional, Tuple
from django.contrib.auth.models import User
from django.core.files.uploadedfile import UploadedFile
from django.utils import timezone
from django.db import transaction
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import Product, Category, ProductImage
from .import_utils import (
    parse_product_sku,
    validate_excel_row,
    get_or_create_category_by_sku,
    process_product_image,
    clean_price_value,
    generate_product_slug,
    ImportStats,
    EXCEL_COLUMN_MAPPING,
    REQUIRED_COLUMNS
)

logger = logging.getLogger(__name__)


class ImportProcessor:
    """
    🔄 Основной класс для обработки импорта товаров из Excel

    ИСПРАВЛЕНО: Устранена ошибка 'list' object is not callable
    """

    def __init__(self, excel_file: UploadedFile, user: User):
        """🚀 Инициализация процессора"""
        self.excel_file = excel_file
        self.user = user
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.stats = ImportStats()
        self.column_mapping = {}
        self.preview_data = []

        logger.info(f"🔄 Инициализирован импорт файла: {excel_file.name} пользователем {user.username}")

    def validate_file(self) -> Tuple[bool, List[str]]:
        """✅ Валидация Excel файла"""
        error_list = []  # 🔧 ИСПРАВЛЕНО: переименовал с errors чтобы избежать конфликта

        try:
            # 📁 Проверка расширения файла
            file_extension = os.path.splitext(self.excel_file.name)[1].lower()
            if file_extension not in ['.xlsx', '.xls']:
                error_list.append(f"Неподдерживаемый формат файла: {file_extension}. Используйте .xlsx или .xls")
                return False, error_list

            # 📊 Проверка размера файла (максимум 10MB)
            if self.excel_file.size > 10 * 1024 * 1024:
                error_list.append(f"Файл слишком большой: {self.excel_file.size / 1024 / 1024:.1f}MB. Максимум: 10MB")
                return False, error_list

            # 📖 Попытка открыть файл
            self.workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
            self.worksheet = self.workbook.active

            if not self.worksheet:
                error_list.append("Не удалось открыть рабочий лист Excel")
                return False, error_list

            # 📊 Проверка наличия данных
            if self.worksheet.max_row < 2:
                error_list.append("Файл должен содержать минимум 2 строки (заголовки + данные)")
                return False, error_list

            # 🔍 Анализ заголовков
            header_validation = self._validate_headers()
            if not header_validation[0]:
                error_list.extend(header_validation[1])
                return False, error_list

            logger.info(f"✅ Файл {self.excel_file.name} прошел валидацию")
            return True, []

        except Exception as e:
            error_msg = f"Ошибка при валидации файла: {str(e)}"
            error_list.append(error_msg)
            logger.error(f"❌ {error_msg}")
            return False, error_list

    def _validate_headers(self) -> Tuple[bool, List[str]]:
        """🔍 Валидация заголовков Excel файла"""
        error_list = []  # 🔧 ИСПРАВЛЕНО: переименовал с errors

        try:
            # 📋 Читаем первую строку (заголовки) - ИСПРАВЛЕНО
            first_row = self.worksheet[1]  # Получаем первую строку
            headers_values = [cell.value for cell in first_row]  # Извлекаем значения

            # 🧹 Очищаем заголовки от None и пробелов
            headers_list = [str(header).strip() if header else f"Колонка_{i + 1}"
                            for i, header in enumerate(headers_values)]

            logger.info(f"📋 Найденные заголовки: {headers_list}")

            # 🎯 Создаем маппинг колонок
            self.column_mapping = {i: header for i, header in enumerate(headers_list)}

            # ✅ Проверка обязательных колонок
            found_required = []
            for required_col in REQUIRED_COLUMNS:
                if required_col in headers_list:
                    found_required.append(required_col)

            missing_required = set(REQUIRED_COLUMNS) - set(found_required)
            if missing_required:
                error_list.append(f"Отсутствуют обязательные колонки: {', '.join(missing_required)}")

            # 📊 Логируем найденные соответствия
            logger.info(f"✅ Найдены обязательные колонки: {found_required}")
            if missing_required:
                logger.warning(f"⚠️ Отсутствуют колонки: {missing_required}")

            return len(missing_required) == 0, error_list

        except Exception as e:
            error_msg = f"Ошибка анализа заголовков: {str(e)}"
            error_list.append(error_msg)
            logger.error(f"❌ {error_msg}")
            return False, error_list

    def preview_data(self, rows_count: int = 5) -> List[Dict[str, Any]]:
        """👁️ Предпросмотр данных для проверки перед импортом"""

        if not self.worksheet:
            logger.error("❌ Рабочий лист не инициализирован")
            return []

        preview_data_list = []  # 🔧 ИСПРАВЛЕНО: переименовал переменную

        try:
            # 📊 Читаем строки данных (начиная со 2-й строки) - ИСПРАВЛЕНО
            max_row_to_read = min(rows_count + 1, self.worksheet.max_row)

            for row_num in range(2, max_row_to_read + 1):  # Начинаем с 2-й строки
                row_data = {}
                current_row = self.worksheet[row_num]  # Получаем строку по номеру

                # 🔄 Преобразуем строку в словарь по заголовкам
                for col_index, cell in enumerate(current_row):
                    column_name = self.column_mapping.get(col_index, f"Колонка_{col_index + 1}")
                    row_data[column_name] = cell.value

                # ➕ Добавляем номер строки для отладки
                row_data['_row_number'] = row_num

                # 🔍 Добавляем результат валидации
                is_valid, validation_errors = validate_excel_row(row_data)
                row_data['_is_valid'] = is_valid
                row_data['_validation_errors'] = validation_errors

                # 📝 Пытаемся распарсить SKU для показа
                product_name = row_data.get('Наименование товара', '')
                category_sku, parsed_name = parse_product_sku(str(product_name))
                row_data['_parsed_category_sku'] = category_sku
                row_data['_parsed_product_name'] = parsed_name

                preview_data_list.append(row_data)

            self.preview_data = preview_data_list
            logger.info(f"👁️ Подготовлен предпросмотр: {len(preview_data_list)} строк")

            return preview_data_list

        except Exception as e:
            logger.error(f"❌ Ошибка предпросмотра данных: {str(e)}")
            return []

    def process_import(self, batch_size: int = 50) -> Dict[str, Any]:
        """🚀 Основная функция импорта данных"""

        if not self.worksheet:
            error_msg = "Рабочий лист не инициализирован. Выполните валидацию файла сначала."
            logger.error(f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}

        # 📊 Инициализация статистики
        self.stats.total_rows = self.worksheet.max_row - 1  # Минус заголовок

        logger.info(f"🚀 Начинаем импорт: {self.stats.total_rows} строк данных")

        try:
            # 🔄 Обработка данных пакетами
            processed_count = 0

            with transaction.atomic():  # 🔒 Используем транзакцию для безопасности

                # 📊 Читаем все строки данных - ИСПРАВЛЕНО
                for row_num in range(2, self.worksheet.max_row + 1):
                    current_row = self.worksheet[row_num]
                    row_values = [cell.value for cell in current_row]  # Извлекаем значения

                    # 🔄 Обработка отдельной строки
                    result = self._process_single_row(row_num, row_values)

                    processed_count += 1
                    self.stats.processed_rows = processed_count

                    # 📊 Логируем прогресс каждые 50 строк
                    if processed_count % batch_size == 0:
                        progress = (processed_count / self.stats.total_rows) * 100
                        logger.info(f"📊 Обработано: {processed_count}/{self.stats.total_rows} ({progress:.1f}%)")

            # 📊 Финальная статистика
            logger.info(f"✅ Импорт завершен: {self.stats.get_summary()}")

            return {
                'success': True,
                'stats': self.stats.get_summary(),
                'message': f"Импорт завершен успешно. Создано: {self.stats.created_count}, обновлено: {self.stats.updated_count}"
            }

        except Exception as e:
            error_msg = f"Критическая ошибка импорта: {str(e)}"
            logger.error(f"❌ {error_msg}")

            return {
                'success': False,
                'error': error_msg,
                'stats': self.stats.get_summary()
            }

    def _process_single_row(self, row_num: int, row_values: List) -> bool:
        """🔄 Обработка одной строки данных - ИСПРАВЛЕНО"""

        try:
            # 🔄 Преобразуем строку в словарь
            row_dict = {}
            for col_index, cell_value in enumerate(row_values):
                column_name = self.column_mapping.get(col_index, f"Колонка_{col_index + 1}")
                row_dict[column_name] = cell_value

            # ✅ Валидация строки
            is_valid, validation_errors = validate_excel_row(row_dict)
            if not is_valid:
                for error in validation_errors:
                    self.stats.add_error(row_num, error)
                return False

            # 🔍 Парсинг SKU
            product_name_raw = row_dict.get('Наименование товара', '')
            category_sku, product_name = parse_product_sku(str(product_name_raw))

            if not category_sku or not product_name:
                self.stats.add_error(row_num, f"Неверный формат названия товара: {product_name_raw}")
                return False

            # 📂 Получение/создание категории
            try:
                category = get_or_create_category_by_sku(category_sku, product_name.split()[0])
            except Exception as e:
                self.stats.add_error(row_num, f"Ошибка работы с категорией: {str(e)}")
                return False

            # 🛍️ Обработка товара
            product_sku = row_dict.get('Код товара', '').strip()
            if not product_sku:
                self.stats.add_error(row_num, "Отсутствует код товара")
                return False

            # 🔍 Поиск существующего товара
            existing_product = Product.objects.filter(product_sku=product_sku).first()

            if existing_product:
                # 🔄 Обновление существующего товара
                updated = self._update_product(existing_product, row_dict, product_name, category)
                if updated:
                    self.stats.add_success(row_num, 'update', existing_product.product_name)

                    # 🖼️ Обработка изображения
                    image_filename = row_dict.get('Изображение')
                    if image_filename:
                        process_product_image(existing_product, str(image_filename))

                return updated
            else:
                # 🆕 Создание нового товара
                new_product = self._create_product(row_dict, product_name, category)
                if new_product:
                    self.stats.add_success(row_num, 'create', new_product.product_name)

                    # 🖼️ Обработка изображения
                    image_filename = row_dict.get('Изображение')
                    if image_filename:
                        process_product_image(new_product, str(image_filename))

                return new_product is not None

        except Exception as e:
            self.stats.add_error(row_num, f"Непредвиденная ошибка: {str(e)}")
            logger.error(f"❌ Ошибка обработки строки {row_num}: {str(e)}")
            return False

    def _create_product(self, row_dict: Dict[str, Any], product_name: str, category: Category) -> Optional[Product]:
        """🆕 Создание нового товара"""

        try:
            # 💰 Обработка цены
            price = clean_price_value(row_dict.get('Цена', 0))

            # 🔗 Генерация slug
            product_sku = row_dict.get('Код товара', '').strip()
            slug = generate_product_slug(product_name, product_sku)

            # 🆕 Создание товара
            product = Product.objects.create(
                product_sku=product_sku,
                product_name=product_name,
                slug=slug,
                category=category,
                price=price,
                product_desription=row_dict.get('Описание товара', ''),
                page_title=row_dict.get('Title страницы', ''),
                meta_description=row_dict.get('Мета-описание', ''),
                newest_product=False  # По умолчанию не новый
            )

            logger.info(f"🆕 Создан товар: {product.product_name} (SKU: {product.product_sku})")
            return product

        except Exception as e:
            logger.error(f"❌ Ошибка создания товара '{product_name}': {str(e)}")
            return None

    def _update_product(self, product: Product, row_dict: Dict[str, Any], product_name: str,
                        category: Category) -> bool:
        """🔄 Обновление существующего товара"""

        try:
            # 💰 Обработка цены
            price = clean_price_value(row_dict.get('Цена', 0))

            # 🔄 Обновление полей
            product.product_name = product_name
            product.category = category
            product.price = price
            product.product_desription = row_dict.get('Описание товара', '')
            product.page_title = row_dict.get('Title страницы', '')
            product.meta_description = row_dict.get('Мета-описание', '')

            # 🔗 Обновляем slug только если он изменился
            new_slug = generate_product_slug(product_name, product.product_sku)
            if product.slug != new_slug:
                product.slug = new_slug

            product.save()

            logger.info(f"🔄 Обновлен товар: {product.product_name} (SKU: {product.product_sku})")
            return True

        except Exception as e:
            logger.error(f"❌ Ошибка обновления товара '{product.product_name}': {str(e)}")
            return False

    def get_file_info(self) -> Dict[str, Any]:
        """📊 Получение информации о файле"""

        if not self.worksheet:
            return {'error': 'Файл не загружен'}

        return {
            'filename': self.excel_file.name,
            'file_size': f"{self.excel_file.size / 1024:.1f} KB",
            'total_rows': self.worksheet.max_row - 1,  # Минус заголовок
            'total_columns': self.worksheet.max_column,
            'column_mapping': self.column_mapping,
            'preview_available': len(self.preview_data) > 0
        }

    def cleanup(self):
        """🧹 Очистка ресурсов"""
        if self.workbook:
            self.workbook.close()
        logger.info("🧹 Ресурсы импорта очищены")

# 🔧 ОСНОВНЫЕ ИСПРАВЛЕНИЯ:
# ✅ ИСПРАВЛЕНО: Переименованы переменные errors -> error_list
# ✅ ИСПРАВЛЕНО: Изменен способ чтения строк Excel без iter_rows
# ✅ ИСПРАВЛЕНО: Убрана передача tuple в _process_single_row
# ✅ ИСПРАВЛЕНО: Прямое обращение к ячейкам worksheet[row_num]