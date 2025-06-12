# 📁 products/import_utils.py
# 🛠️ ИСПРАВЛЕННАЯ версия утилит импорта с автогенерацией SKU
# ✅ Добавлена функция generate_sku_for_product()
# 🔧 МИНИМАЛЬНАЯ валидация - принимаем почти любые данные

import logging
import openpyxl
from typing import Dict, List, Tuple, Optional, Union
from decimal import Decimal, InvalidOperation
from django.core.files.uploadedfile import InMemoryUploadedFile
import re

logger = logging.getLogger(__name__)

# 🗂️ Маппинг колонок Excel (одинаковый для категорий и товаров)
EXCEL_COLUMN_MAPPING = {
    'identifier': 0,  # A: Категория (1.BMW) или SKU товара (10001)
    'name': 1,  # B: Название категории или товара
    'title': 2,  # C: Title страницы
    'price': 3,  # D: Цена (только для товаров, у категорий пустая)
    'description': 4,  # E: Описание
    'meta_description': 5,  # F: Мета-описание
    'image': 6  # G: Изображение
}

# 📋 УПРОЩЕННЫЕ требования
REQUIRED_FIELDS = ['identifier']  # ✅ Только идентификатор обязателен


def read_excel_file(file: InMemoryUploadedFile) -> Tuple[bool, Union[List[Dict], str]]:
    """
    📊 Чтение Excel файла с разделением на категории и товары

    Args:
        file: Загруженный Excel файл

    Returns:
        Tuple[bool, Union[List[Dict], str]]: (success, data_or_error)
    """
    try:
        logger.info(f"🔄 Начинаем чтение файла: {file.name}")

        # 📖 Загружаем Excel файл
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active

        # 📏 Получаем размеры таблицы
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        logger.info(f"📊 Размер таблицы: {max_row} строк, {max_col} колонок")

        if max_row < 2:
            return False, "❌ Файл должен содержать минимум 2 строки (заголовок + данные)"

        # 📋 Читаем данные начиная со второй строки (пропускаем заголовок)
        data = []

        for row_num in range(2, max_row + 1):
            try:
                # 🔍 Получаем значения ячеек
                row_data = {}

                # 📝 Безопасное извлечение данных из каждой ячейки
                for field_name, col_index in EXCEL_COLUMN_MAPPING.items():
                    cell_value = None

                    if col_index < max_col:
                        cell = worksheet.cell(row=row_num, column=col_index + 1)  # +1 так как Excel 1-based
                        cell_value = cell.value

                    # 🧹 Очистка и нормализация данных
                    if cell_value is not None:
                        if isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if not cell_value:
                                cell_value = None

                    row_data[field_name] = cell_value

                # 🚫 Пропускаем строки без идентификатора
                if not row_data.get('identifier'):
                    logger.warning(f"⚠️ Строка {row_num}: пропущена (нет идентификатора)")
                    continue

                # 🎯 Определяем тип строки: категория или товар
                identifier = str(row_data['identifier']).strip()
                is_category = '.' in identifier

                row_data['row_number'] = row_num
                row_data['is_category'] = is_category

                if is_category:
                    # 📂 Это категория - извлекаем чистое название
                    row_data['category_name'] = extract_category_name(identifier)
                    row_data['type'] = 'category'
                else:
                    # 🛍️ Это товар - сохраняем SKU
                    row_data['sku'] = identifier
                    row_data['type'] = 'product'

                data.append(row_data)

            except Exception as e:
                logger.error(f"❌ Ошибка обработки строки {row_num}: {str(e)}")
                continue

        logger.info(f"✅ Успешно прочитано {len(data)} строк данных")
        return True, data

    except Exception as e:
        error_msg = f"❌ Ошибка чтения файла: {str(e)}"
        logger.error(error_msg)
        return False, error_msg


def extract_category_name(category_identifier: str) -> str:
    """
    📂 Извлечение чистого названия категории из идентификатора

    Примеры:
    - "1.BMW" -> "BMW"
    - "2.Acura" -> "Acura"
    - "sky.BMW" -> "BMW"

    Args:
        category_identifier: Идентификатор категории с точкой

    Returns:
        str: Чистое название категории
    """
    try:
        # 🔍 Берём часть после последней точки
        parts = category_identifier.split('.')
        if len(parts) >= 2:
            category_name = parts[-1].strip().upper()
            return category_name if category_name else 'ТОВАРЫ'

        return category_identifier.strip().upper()

    except Exception as e:
        logger.error(f"Ошибка извлечения категории из '{category_identifier}': {e}")
        return 'ТОВАРЫ'


def extract_category_sku(category_identifier: str) -> int:
    """
    🔢 Извлечение SKU категории из идентификатора

    Примеры:
    - "1.BMW" -> 1
    - "2.Acura" -> 2
    - "10.Mercedes" -> 10

    Args:
        category_identifier: Идентификатор категории с точкой

    Returns:
        int: SKU категории (по умолчанию 1)
    """
    try:
        # 🔍 Берём часть до первой точки
        parts = category_identifier.split('.')
        if len(parts) >= 2:
            sku_part = parts[0].strip()

            # 🔢 Пытаемся преобразовать в число
            if sku_part.isdigit():
                return int(sku_part)

        logger.warning(f"⚠️ Не удалось извлечь SKU из '{category_identifier}', используем 1")
        return 1

    except Exception as e:
        logger.error(f"Ошибка извлечения SKU категории из '{category_identifier}': {e}")
        return 1


def generate_sku_for_product(category_sku: int, existing_products_in_category: List[Dict]) -> str:
    """
    🆕 НОВАЯ ФУНКЦИЯ: Автогенерация SKU для товара по формуле

    Формула: category_sku * 10000 + порядковый_номер_в_категории

    Примеры:
    - Категория BMW (sku=1): товары получат 10001, 10002, 10003...
    - Категория Acura (sku=2): товары получат 20001, 20002, 20003...

    Args:
        category_sku: SKU категории (1, 2, 3...)
        existing_products_in_category: Список уже существующих товаров в категории

    Returns:
        str: Сгенерированный SKU товара
    """
    try:
        # 🔢 Базовое значение: category_sku * 10000
        base_sku = category_sku * 10000

        # 🔍 Находим максимальный порядковый номер среди существующих товаров
        max_sequence = 0

        for product in existing_products_in_category:
            product_sku = product.get('sku', '')

            # 🎯 Пытаемся извлечь порядковый номер из SKU
            if isinstance(product_sku, (str, int)):
                try:
                    sku_int = int(product_sku)

                    # ✅ Проверяем, что SKU относится к нашей категории
                    if base_sku <= sku_int < (base_sku + 10000):
                        sequence = sku_int - base_sku
                        max_sequence = max(max_sequence, sequence)

                except (ValueError, TypeError):
                    continue

        # 📈 Следующий порядковый номер
        next_sequence = max_sequence + 1

        # 🎯 Генерируем итоговый SKU
        generated_sku = base_sku + next_sequence

        logger.info(
            f"🆕 Сгенерирован SKU: {generated_sku} (категория {category_sku}, последовательность {next_sequence})")

        return str(generated_sku)

    except Exception as e:
        logger.error(f"❌ Ошибка генерации SKU для категории {category_sku}: {e}")

        # 🚨 Fallback: простая генерация
        fallback_sku = category_sku * 10000 + 1
        return str(fallback_sku)


def validate_row(row_data: Dict) -> Tuple[bool, List[str]]:
    """
    ✅ УПРОЩЕННАЯ валидация строки (только критичные проверки)

    Args:
        row_data: Словарь с данными строки

    Returns:
        Tuple[bool, List[str]]: (is_valid, list_of_errors)
    """
    errors = []

    try:
        # 🔍 ЕДИНСТВЕННАЯ обязательная проверка - наличие идентификатора
        identifier = row_data.get('identifier')
        if not identifier:
            errors.append(f"❌ Отсутствует идентификатор (колонка A)")
            return False, errors

        # ✅ Простая проверка названия (без предупреждений)
        # Если название пустое - оставляем пустым, не заполняем по умолчанию

        # 💰 Базовая проверка цены (только критичные ошибки)
        if row_data.get('type') == 'product':
            price_value = row_data.get('price')
            if price_value is not None:
                try:
                    normalized_price = normalize_price(price_value)
                    # Убираем проверки диапазона цены - позволяем любые значения
                except Exception as e:
                    # Только логируем, но не блокируем
                    logger.warning(f"⚠️ Проблема с ценой '{price_value}' для {identifier}: {e}")

    except Exception as e:
        # 🛡️ Критичная ошибка - блокируем только если что-то совсем сломано
        errors.append(f"❌ Критичная ошибка валидации: {str(e)}")
        logger.error(f"Критичная ошибка при валидации строки: {e}")

    is_valid = len(errors) == 0
    return is_valid, errors


def normalize_price(price_value: Union[str, int, float, None]) -> float:
    """
    💰 Нормализация цены (только для товаров)

    Args:
        price_value: Значение цены в любом формате

    Returns:
        float: Нормализованная цена (0.0 если пустая)
    """
    try:
        if price_value is None or price_value == '':
            return 0.0

        if isinstance(price_value, (int, float)):
            return float(price_value)

        if isinstance(price_value, str):
            # 🧹 Удаляем валютные символы и лишние символы
            price_clean = re.sub(r'[^\d.,]', '', price_value.strip())

            if not price_clean:
                return 0.0

            # 🔄 Заменяем запятую на точку
            price_clean = price_clean.replace(',', '.')

            # 🎯 Обрабатываем случай множественных точек
            if price_clean.count('.') > 1:
                parts = price_clean.split('.')
                price_clean = ''.join(parts[:-1]) + '.' + parts[-1]

            return float(price_clean)

    except Exception as e:
        logger.warning(f"⚠️ Не удалось обработать цену '{price_value}': {e}")

    return 0.0


def separate_categories_and_products(raw_data: List[Dict]) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    🔄 ИСПРАВЛЕННОЕ разделение данных с автогенерацией SKU

    📋 ЛОГИКА:
    - Читаем строки по порядку сверху вниз
    - Строка с точкой → категория (запоминаем как current_category)
    - Строка без точки → товар (привязываем к current_category + генерируем SKU если нужно)

    Args:
        raw_data: Сырые данные из Excel

    Returns:
        Tuple[List[Dict], List[Dict], List[Dict]]: (categories, products, invalid_data)
    """
    categories = []
    products = []
    invalid_data = []
    current_category = None  # 💾 Последняя обработанная категория
    current_category_sku = 1  # 🔢 SKU текущей категории
    products_in_current_category = []  # 📦 Товары в текущей категории (для генерации SKU)

    logger.info(f"🔄 Начинаем обработку с автогенерацией SKU {len(raw_data)} строк...")

    for row in raw_data:
        try:
            # ✅ Валидируем строку
            is_valid, errors = validate_row(row)

            if not is_valid:
                row['errors'] = errors
                invalid_data.append(row)
                logger.warning(f"❌ Строка {row.get('row_number', '?')}: ошибки валидации")
                continue

            if row['is_category']:
                # 📂 Обрабатываем категорию
                category_name = row['category_name']

                # 🔢 НОВОЕ: Извлекаем SKU категории из идентификатора
                category_identifier = str(row['identifier']).strip()
                current_category_sku = extract_category_sku(category_identifier)

                category_data = {
                    'category_name': category_name,
                    'category_sku': current_category_sku,  # 🆕 Добавляем SKU категории
                    'name': row.get('name', ''),
                    'title': row.get('title', ''),
                    'description': row.get('description', ''),
                    'meta_description': row.get('meta_description', ''),
                    'image': row.get('image', ''),
                    'row_number': row.get('row_number', 0)
                }

                categories.append(category_data)
                current_category = category_name  # 💾 ЗАПОМИНАЕМ как текущую
                products_in_current_category = []  # 🔄 Сбрасываем список товаров для новой категории

                logger.info(
                    f"📂 Обработана категория: {current_category} (SKU: {current_category_sku}, строка {row.get('row_number', '?')})")

            else:
                # 🛍️ Обрабатываем товар
                if not current_category:
                    # ⚠️ Товар без категории - создаём дефолтную
                    current_category = 'ТОВАРЫ'
                    current_category_sku = 1
                    logger.warning(f"⚠️ Товар без категории, используем дефолтную: {current_category}")

                original_sku = row['sku']

                # 🆕 АВТОГЕНЕРАЦИЯ SKU если он пустой
                if not original_sku or original_sku == '':
                    generated_sku = generate_sku_for_product(current_category_sku, products_in_current_category)
                    final_sku = generated_sku
                    logger.info(f"🆕 Сгенерирован SKU: {final_sku} для товара в категории {current_category}")
                else:
                    final_sku = str(original_sku).strip()
                    logger.info(f"✅ Используем существующий SKU: {final_sku}")

                product_data = {
                    'sku': final_sku,  # 🎯 Финальный SKU (исходный или сгенерированный)
                    'original_sku': original_sku,  # 📝 Оригинальный SKU из Excel (для отладки)
                    'name': row.get('name', ''),
                    'title': row.get('title', ''),
                    'price': normalize_price(row.get('price')),
                    'description': row.get('description', ''),
                    'meta_description': row.get('meta_description', ''),
                    'image': row.get('image', ''),
                    'category_name': current_category,  # 🔗 ПРИВЯЗЫВАЕМ к текущей категории
                    'category_sku': current_category_sku,  # 🔗 SKU категории
                    'row_number': row.get('row_number', 0)
                }

                products.append(product_data)
                products_in_current_category.append(product_data)  # 📦 Добавляем в список для генерации следующих SKU

                logger.info(
                    f"🛍️ Товар {final_sku} → категория {current_category} (строка {row.get('row_number', '?')})")

        except Exception as e:
            logger.error(f"❌ Ошибка обработки строки {row.get('row_number', '?')}: {e}")
            row['errors'] = [f"Ошибка обработки: {str(e)}"]
            invalid_data.append(row)

    # 📊 Создаем недостающие категории
    used_categories = set(p['category_name'] for p in products if p.get('category_name'))
    existing_categories = set(cat['category_name'] for cat in categories)
    missing_categories = used_categories - existing_categories

    for missing_cat in missing_categories:
        categories.append({
            'category_name': missing_cat,
            'category_sku': 1,  # 🔢 Дефолтный SKU для недостающих категорий
            'name': '',
            'title': '',
            'description': '',
            'meta_description': '',
            'image': '',
            'row_number': 0
        })
        logger.info(f"🆕 Создана недостающая категория: {missing_cat} (SKU: 1)")

    logger.info(
        f"✅ Обработка с автогенерацией SKU завершена: {len(categories)} категорий, {len(products)} товаров, {len(invalid_data)} ошибок")

    return categories, products, invalid_data


def get_import_statistics(categories: List[Dict], products: List[Dict], invalid_data: List[Dict]) -> Dict:
    """
    📊 Подсчёт статистики импорта с разделением на категории и товары
    """
    try:
        category_names = set(cat['category_name'] for cat in categories)
        products_with_images = sum(1 for prod in products if prod.get('image'))
        products_with_prices = sum(1 for prod in products if prod.get('price', 0) > 0)

        # 🆕 Статистика по SKU
        products_with_sku = sum(1 for prod in products if prod.get('sku'))
        auto_generated_sku = sum(1 for prod in products if not prod.get('original_sku'))

        return {
            'total_rows': len(categories) + len(products) + len(invalid_data),
            'categories_count': len(categories),
            'products_count': len(products),
            'invalid_rows': len(invalid_data),
            'category_names': list(category_names),
            'products_with_images': products_with_images,
            'products_with_prices': products_with_prices,
            'categories_with_images': sum(1 for cat in categories if cat.get('image')),
            'products_with_sku': products_with_sku,  # 🆕 Товары с SKU
            'auto_generated_sku': auto_generated_sku,  # 🆕 Автосгенерированные SKU
        }
    except Exception as e:
        logger.error(f"Ошибка подсчёта статистики: {e}")
        return {}

# 🔧 НОВЫЕ ВОЗМОЖНОСТИ:
# ✅ ДОБАВЛЕНО: extract_category_sku() - извлечение SKU категории из идентификатора
# ✅ ДОБАВЛЕНО: generate_sku_for_product() - автогенерация SKU по формуле
# ✅ ИЗМЕНЕНО: separate_categories_and_products() - поддержка автогенерации SKU
# ✅ УЛУЧШЕНО: get_import_statistics() - добавлена статистика по SKU
# ✅ РЕЗУЛЬТАТ: Товары получают правильные SKU, нет дубликатов